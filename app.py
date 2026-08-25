from pathlib import Path
from openpyxl import load_workbook
from dotenv import load_dotenv
from groq import Groq
from pypdf import PdfReader

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import os


# =========================================================
# PAGE CONFIG
# =========================================================

st.set_page_config(
    page_title="HUL AI Financial Analyst",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)


# =========================================================
# DESIGN
# =========================================================

PRIMARY = "#2563EB"
NAVY = "#0F172A"
TEXT = "#334155"
MUTED = "#64748B"
BORDER = "#E2E8F0"
BG = "#F8FAFC"
WHITE = "#FFFFFF"


st.html(
    f"""
    <style>

    .stApp {{
        background: {BG};
    }}

    .block-container {{
        max-width: 1380px;
        padding-top: 2rem;
        padding-bottom: 3rem;
    }}

    #MainMenu {{
        visibility: hidden;
    }}

    footer {{
        visibility: hidden;
    }}

    header[data-testid="stHeader"] {{
        background: transparent;
    }}

    h1, h2, h3 {{
        color: {NAVY};
    }}

    .hero {{
        background: {WHITE};
        border: 1px solid {BORDER};
        border-left: 5px solid {PRIMARY};
        padding: 25px 28px;
        border-radius: 5px;
        margin-bottom: 22px;
    }}

    .hero-title {{
        color: {NAVY};
        font-size: 30px;
        font-weight: 750;
        margin-bottom: 6px;
    }}

    .hero-subtitle {{
        color: {MUTED};
        font-size: 14px;
    }}

    .kpi-card {{
        background: {WHITE};
        border: 1px solid {BORDER};
        border-radius: 5px;
        padding: 18px 20px;
        min-height: 120px;
    }}

    .kpi-label {{
        color: {MUTED};
        font-size: 12px;
        font-weight: 700;
        letter-spacing: .04em;
        margin-bottom: 9px;
    }}

    .kpi-value {{
        color: {NAVY};
        font-size: 25px;
        font-weight: 750;
    }}

    .kpi-note {{
        color: {MUTED};
        font-size: 12px;
        margin-top: 9px;
    }}

    .section {{
        margin-top: 10px;
        margin-bottom: 18px;
    }}

    .eyebrow {{
        color: {PRIMARY};
        font-size: 11px;
        font-weight: 750;
        text-transform: uppercase;
        letter-spacing: .08em;
    }}

    .section-title {{
        color: {NAVY};
        font-size: 23px;
        font-weight: 750;
        margin-top: 3px;
    }}

    .section-desc {{
        color: {MUTED};
        font-size: 13px;
        margin-top: 5px;
    }}

    .note {{
        background: #EFF6FF;
        border-left: 4px solid {PRIMARY};
        color: {TEXT};
        padding: 13px 16px;
        border-radius: 4px;
        font-size: 13px;
        margin: 10px 0 20px 0;
    }}

    .source-card {{
        background: {WHITE};
        border: 1px solid {BORDER};
        padding: 15px;
        border-radius: 5px;
        min-height: 120px;
        color: {TEXT};
        font-size: 13px;
    }}

    .stTabs [data-baseweb="tab-list"] {{
        background: {WHITE};
        border: 1px solid {BORDER};
        border-radius: 5px;
        padding: 5px;
        gap: 4px;
    }}

    .stTabs [data-baseweb="tab"] {{
        height: 42px;
        padding-left: 14px;
        padding-right: 14px;
        color: {MUTED};
        font-weight: 600;
        font-size: 13px;
    }}

    .stTabs [aria-selected="true"] {{
        background: #EFF6FF !important;
        color: {PRIMARY} !important;
        border-radius: 4px;
    }}

    .stButton > button {{
        background: {PRIMARY};
        color: white;
        border: 1px solid {PRIMARY};
        border-radius: 4px;
        font-weight: 650;
        min-height: 42px;
        padding: 0 18px;
    }}

    .stButton > button:hover {{
        background: {NAVY};
        border-color: {NAVY};
        color: white;
    }}

    div[data-testid="stMetric"] {{
        background: {WHITE};
        border: 1px solid {BORDER};
        border-radius: 5px;
        padding: 17px;
    }}

    div[data-testid="stMetricLabel"] {{
        color: {MUTED};
        font-weight: 650;
    }}

    div[data-testid="stMetricValue"] {{
        color: {NAVY};
        font-weight: 750;
    }}

    .footer-box {{
        background: {WHITE};
        border: 1px solid {BORDER};
        border-top: 3px solid {PRIMARY};
        padding: 18px 20px;
        border-radius: 5px;
        color: {MUTED};
        font-size: 12px;
        line-height: 1.7;
        margin-top: 35px;
    }}

    </style>
    """
)


# =========================================================
# ENVIRONMENT
# =========================================================

load_dotenv()
api_key = os.getenv("GROQ_API_KEY")


# =========================================================
# FIND EXCEL
# =========================================================

def find_financial_workbook():

    for excel_file in Path(".").glob("*.xlsx"):

        if excel_file.name.startswith("~$"):
            continue

        try:

            wb = load_workbook(
                excel_file,
                data_only=True,
                read_only=True
            )

            sheets = wb.sheetnames
            wb.close()

            if "Raw Data" in sheets and "Ratios" in sheets:
                return excel_file

        except Exception:
            continue

    return None


file_path = find_financial_workbook()


if file_path is None:

    st.error(
        "HUL Excel workbook not found. "
        "Workbook must contain Raw Data and Ratios sheets."
    )

    st.stop()


# =========================================================
# LOAD EXCEL
# =========================================================

workbook = load_workbook(
    file_path,
    data_only=True
)

raw = workbook["Raw Data"]
ratios = workbook["Ratios"]


def get_ratio(name, column):

    for row in range(
        1,
        ratios.max_row + 1
    ):

        value = ratios.cell(
            row=row,
            column=1
        ).value

        if value:

            if (
                str(value).strip().lower()
                == name.strip().lower()
            ):

                return ratios[f"{column}{row}"].value

    return None


# =========================================================
# FORMAT FUNCTIONS
# =========================================================

def format_percent(value, decimals=2):

    if value is None:
        return "N/A"

    return f"{value * 100:.{decimals}f}%"


def format_money(value):

    if value is None:
        return "N/A"

    return f"₹{value:,.0f} Cr"


def format_number(value, decimals=1):

    if value is None:
        return "N/A"

    return f"{value:.{decimals}f}"


# =========================================================
# FY25 FINANCIAL VALUES
# =========================================================

revenue = raw["F4"].value
ebitda = raw["F5"].value
pat = raw["F6"].value

cash = raw["F16"].value or 0

non_current_lease = raw["F17"].value or 0
current_lease = raw["F18"].value or 0

lease_liabilities = (
    non_current_lease
    + current_lease
)


ebitda_margin = get_ratio(
    "EBITDA Margin",
    "F"
)

roce = get_ratio(
    "ROCE (Reported)",
    "F"
)

roe = get_ratio(
    "ROE",
    "F"
)

revenue_cagr = get_ratio(
    "Revenue CAGR",
    "B"
)

ebitda_cagr = get_ratio(
    "EBITDA CAGR",
    "B"
)

pat_cagr = get_ratio(
    "PAT CAGR",
    "B"
)

fy25_revenue_growth = get_ratio(
    "Revenue YoY Growth",
    "F"
)

fy25_ebitda_growth = get_ratio(
    "EBITDA YoY Growth",
    "F"
)

fy25_pat_growth = get_ratio(
    "PAT YoY Growth",
    "F"
)

ccc = get_ratio(
    "Cash Conversion Cycle",
    "F"
)


# =========================================================
# HISTORICAL DATA
# =========================================================

years = [
    raw[f"{c}3"].value
    for c in ["B", "C", "D", "E", "F"]
]

revenue_values = [
    raw[f"{c}4"].value
    for c in ["B", "C", "D", "E", "F"]
]

ebitda_values = [
    raw[f"{c}5"].value
    for c in ["B", "C", "D", "E", "F"]
]

pat_values = [
    raw[f"{c}6"].value
    for c in ["B", "C", "D", "E", "F"]
]


# =========================================================
# PDF LOADER
# =========================================================

@st.cache_data(show_spinner=False)
def load_pdf_pages():

    pages = []

    for pdf_file in Path(".").glob("*.pdf"):

        try:

            reader = PdfReader(
                str(pdf_file)
            )

            for page_number, page in enumerate(
                reader.pages,
                start=1
            ):

                text = page.extract_text() or ""

                if text.strip():

                    pages.append(
                        {
                            "file": pdf_file.name,
                            "page": page_number,
                            "text": text
                        }
                    )

        except Exception:
            continue

    return pages


# =========================================================
# PDF SEARCH
# =========================================================

def search_annual_reports(
    question,
    pdf_pages,
    top_k=3
):

    question_lower = question.lower()

    stop_words = {
        "what", "where", "when", "which",
        "with", "from", "that", "this",
        "does", "have", "about",
        "financial", "company", "hul"
    }

    words = []

    for word in question_lower.split():

        clean_word = word.strip(
            ".,?!'\"()"
        )

        if (
            len(clean_word) > 3
            and clean_word not in stop_words
        ):

            words.append(clean_word)


    extra_terms = []


    if "roce" in question_lower:

        extra_terms.extend(
            [
                "return on capital employed",
                "capital employed"
            ]
        )


    if "roe" in question_lower:

        extra_terms.append(
            "return on equity"
        )


    if "debt" in question_lower:

        extra_terms.extend(
            [
                "borrowings",
                "debt equity",
                "lease liabilities"
            ]
        )


    if "working capital" in question_lower:

        extra_terms.extend(
            [
                "working capital",
                "current assets",
                "current liabilities"
            ]
        )


    if "ebitda" in question_lower:

        extra_terms.append(
            "ebitda"
        )


    results = []


    for page in pdf_pages:

        text_lower = page["text"].lower()

        score = 0


        for word in set(words):

            score += text_lower.count(
                word
            )


        for term in extra_terms:

            if term in text_lower:
                score += 5


        if score > 0:

            results.append(
                {
                    **page,
                    "score": score
                }
            )


    results.sort(
        key=lambda x: x["score"],
        reverse=True
    )


    return results[:top_k]


# =========================================================
# FORECAST FUNCTION
# =========================================================

def create_forecast(
    starting_revenue,
    growth,
    ebitda_margin_value,
    pat_margin_value
):

    forecast_years = [
        "FY26",
        "FY27",
        "FY28"
    ]

    revenues = []
    ebitdas = []
    pats = []

    current_revenue = starting_revenue


    for _ in forecast_years:

        current_revenue = (
            current_revenue
            * (1 + growth / 100)
        )

        future_ebitda = (
            current_revenue
            * ebitda_margin_value
            / 100
        )

        future_pat = (
            current_revenue
            * pat_margin_value
            / 100
        )

        revenues.append(current_revenue)
        ebitdas.append(future_ebitda)
        pats.append(future_pat)


    return pd.DataFrame(
        {
            "Revenue": revenues,
            "EBITDA": ebitdas,
            "PAT": pats
        },
        index=forecast_years
    )


# =========================================================
# CHART STYLE
# =========================================================

def style_chart(fig, height=420):

    fig.update_layout(
        height=height,
        paper_bgcolor=WHITE,
        plot_bgcolor=WHITE,
        font=dict(
            family="Arial",
            size=12,
            color=TEXT
        ),
        margin=dict(
            l=35,
            r=25,
            t=65,
            b=40
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.03,
            xanchor="left",
            x=0
        ),
        hoverlabel=dict(
            bgcolor=WHITE,
            font_size=12
        )
    )

    fig.update_xaxes(
        showgrid=False,
        linecolor=BORDER
    )

    fig.update_yaxes(
        gridcolor="#EDF2F7",
        zeroline=False
    )


# =========================================================
# HERO
# =========================================================

st.html(
    """
    <div class="hero">

        <div class="hero-title">
            HUL AI Financial Analyst
        </div>

        <div class="hero-subtitle">
            Hindustan Unilever Limited ·
            5-Year Financial Analysis ·
            Excel Financial Model ·
            AI Intelligence ·
            Annual Report Research ·
            Forecasting ·
            Valuation
        </div>

    </div>
    """
)


# =========================================================
# KPI CARDS
# =========================================================

k1, k2, k3, k4 = st.columns(4)


with k1:

    st.html(
        f"""
        <div class="kpi-card">

            <div class="kpi-label">
                FY25 REVENUE
            </div>

            <div class="kpi-value">
                {format_money(revenue)}
            </div>

            <div class="kpi-note">
                YoY Growth:
                {format_percent(fy25_revenue_growth)}
            </div>

        </div>
        """
    )


with k2:

    st.html(
        f"""
        <div class="kpi-card">

            <div class="kpi-label">
                EBITDA MARGIN
            </div>

            <div class="kpi-value">
                {format_percent(ebitda_margin)}
            </div>

            <div class="kpi-note">
                EBITDA:
                {format_money(ebitda)}
            </div>

        </div>
        """
    )


with k3:

    st.html(
        f"""
        <div class="kpi-card">

            <div class="kpi-label">
                FY25 PAT
            </div>

            <div class="kpi-value">
                {format_money(pat)}
            </div>

            <div class="kpi-note">
                YoY Growth:
                {format_percent(fy25_pat_growth)}
            </div>

        </div>
        """
    )


with k4:

    st.html(
        f"""
        <div class="kpi-card">

            <div class="kpi-label">
                RETURN ON CAPITAL
            </div>

            <div class="kpi-value">
                {format_percent(roce, 1)}
            </div>

            <div class="kpi-note">
                FY25 ROE:
                {format_percent(roe)}
            </div>

        </div>
        """
    )


st.write("")


# =========================================================
# TABS
# =========================================================

(
    tab1,
    tab2,
    tab3,
    tab4,
    tab5,
    tab6,
    tab7
) = st.tabs(
    [
        "Overview",
        "AI Analysis",
        "Annual Report Q&A",
        "Risk & Opportunity",
        "Forecast",
        "Sensitivity",
        "Valuation"
    ]
)


# =========================================================
# TAB 1 — OVERVIEW
# =========================================================

with tab1:

    st.html(
        """
        <div class="section">

            <div class="eyebrow">
                Historical Performance
            </div>

            <div class="section-title">
                5-Year Financial Overview
            </div>

            <div class="section-desc">
                Revenue, EBITDA and PAT performance
                from FY2020-21 to FY2024-25.
            </div>

        </div>
        """
    )


    historical_fig = go.Figure()


    historical_fig.add_trace(
        go.Scatter(
            x=years,
            y=revenue_values,
            mode="lines+markers",
            name="Revenue",
            line=dict(
                color=PRIMARY,
                width=3
            ),
            marker=dict(size=8),
            hovertemplate=(
                "<b>Revenue</b><br>"
                "%{x}<br>"
                "₹%{y:,.0f} Cr"
                "<extra></extra>"
            )
        )
    )


    historical_fig.add_trace(
        go.Scatter(
            x=years,
            y=ebitda_values,
            mode="lines+markers",
            name="EBITDA",
            line=dict(
                color="#475569",
                width=2.5
            ),
            marker=dict(size=7),
            hovertemplate=(
                "<b>EBITDA</b><br>"
                "%{x}<br>"
                "₹%{y:,.0f} Cr"
                "<extra></extra>"
            )
        )
    )


    historical_fig.add_trace(
        go.Scatter(
            x=years,
            y=pat_values,
            mode="lines+markers",
            name="PAT",
            line=dict(
                color="#94A3B8",
                width=2.5
            ),
            marker=dict(size=7),
            hovertemplate=(
                "<b>PAT</b><br>"
                "%{x}<br>"
                "₹%{y:,.0f} Cr"
                "<extra></extra>"
            )
        )
    )


    historical_fig.update_layout(
        title="Revenue, EBITDA & PAT Trend",
        yaxis_title="₹ Crore"
    )


    style_chart(
        historical_fig
    )


    st.plotly_chart(
        historical_fig,
        use_container_width=True,
        config={
            "displayModeBar": False
        }
    )


    st.subheader(
        "Long-Term Growth"
    )


    g1, g2, g3 = st.columns(3)


    g1.metric(
        "Revenue CAGR",
        format_percent(
            revenue_cagr
        )
    )


    g2.metric(
        "EBITDA CAGR",
        format_percent(
            ebitda_cagr
        )
    )


    g3.metric(
        "PAT CAGR",
        format_percent(
            pat_cagr
        )
    )


    st.html(
        """
        <div class="note">

            <b>Period:</b>
            FY2020-21 to FY2024-25

            &nbsp;&nbsp; | &nbsp;&nbsp;

            <b>Basis:</b>
            Standalone

            &nbsp;&nbsp; | &nbsp;&nbsp;

            <b>Currency:</b>
            ₹ Crore

        </div>
        """
    )


# =========================================================
# TAB 2 — AI ANALYSIS
# =========================================================

with tab2:

    st.html(
        """
        <div class="section">

            <div class="eyebrow">
                AI Interpretation
            </div>

            <div class="section-title">
                Financial Performance Analysis
            </div>

            <div class="section-desc">
                AI interprets calculated financial metrics
                from the Excel model.
            </div>

        </div>
        """
    )


    a1, a2, a3 = st.columns(3)


    a1.metric(
        "Revenue Growth",
        format_percent(
            fy25_revenue_growth
        )
    )


    a2.metric(
        "EBITDA Growth",
        format_percent(
            fy25_ebitda_growth
        )
    )


    a3.metric(
        "PAT Growth",
        format_percent(
            fy25_pat_growth
        )
    )


    st.write("")


    if st.button(
        "Generate AI Analysis",
        key="analysis_button"
    ):

        if not api_key:

            st.error(
                "Groq API key not found."
            )

        else:

            analysis_prompt = f"""
You are a professional financial analyst.

Analyse Hindustan Unilever Limited
using only the supplied financial metrics.

Analysis Period:
FY2020-21 to FY2024-25.

FY25 Revenue:
{format_money(revenue)}

Revenue Growth:
{format_percent(fy25_revenue_growth)}

EBITDA:
{format_money(ebitda)}

EBITDA Margin:
{format_percent(ebitda_margin)}

EBITDA Growth:
{format_percent(fy25_ebitda_growth)}

FY25 PAT:
{format_money(pat)}

PAT Growth:
{format_percent(fy25_pat_growth)}

ROE:
{format_percent(roe)}

ROCE:
{format_percent(roce, 1)}

Revenue CAGR:
{format_percent(revenue_cagr)}

EBITDA CAGR:
{format_percent(ebitda_cagr)}

PAT CAGR:
{format_percent(pat_cagr)}

Cash Conversion Cycle:
{format_number(ccc)} days


Give exactly:

### Growth
### Profitability
### Return Efficiency
### Working Capital
### Key Strength
### Key Concern
### Overall View


Rules:

- Use only supplied numbers.
- Do not invent business causes.
- Do not use LaTeX.
- Keep analysis concise.
"""


            client = Groq(
                api_key=api_key
            )


            try:

                with st.spinner(
                    "Generating financial analysis..."
                ):

                    response = (
                        client.chat.completions.create(
                            model="openai/gpt-oss-20b",
                            messages=[
                                {
                                    "role": "user",
                                    "content": analysis_prompt
                                }
                            ]
                        )
                    )


                with st.container(
                    border=True
                ):

                    st.markdown(
                        response
                        .choices[0]
                        .message
                        .content
                    )


            except Exception as error:

                st.error(
                    f"AI request failed: {error}"
                )


# =========================================================
# TAB 3 — ANNUAL REPORT Q&A
# =========================================================

with tab3:

    st.html(
        """
        <div class="section">

            <div class="eyebrow">
                Source-Grounded Research
            </div>

            <div class="section-title">
                Ask HUL Annual Reports
            </div>

            <div class="section-desc">
                Search Annual Reports and receive answers
                with supporting PDF page references.
            </div>

        </div>
        """
    )


    st.html(
        """
        <div class="note">

            Try:
            <b>
            What is HUL's ROCE and how is
            capital employed calculated?
            </b>

        </div>
        """
    )


    question = st.text_input(
        "Financial question",
        placeholder="Enter your question...",
        key="annual_report_question"
    )


    if st.button(
        "Search & Ask AI",
        key="ask_ai_button"
    ):

        if not question:

            st.warning(
                "Please enter a question."
            )


        elif not api_key:

            st.error(
                "Groq API key not found."
            )


        else:

            with st.spinner(
                "Searching HUL Annual Reports..."
            ):

                pdf_pages = load_pdf_pages()

                matches = search_annual_reports(
                    question,
                    pdf_pages,
                    top_k=3
                )


            source_context = ""


            for index, result in enumerate(
                matches,
                start=1
            ):

                source_context += f"""
[Source {index}]

File:
{result['file']}

PDF Page:
{result['page']}

Annual Report Text:

{result['text'][:2500]}

"""


            question_prompt = f"""
You are a professional financial analyst.

Answer the user's question using ONLY:

1. Excel financial data.
2. Retrieved HUL Annual Report text.


EXCEL DATA

FY25 Revenue:
{format_money(revenue)}

EBITDA:
{format_money(ebitda)}

EBITDA Margin:
{format_percent(ebitda_margin)}

FY25 PAT:
{format_money(pat)}

ROE:
{format_percent(roe)}

ROCE:
{format_percent(roce, 1)}

Revenue CAGR:
{format_percent(revenue_cagr)}

EBITDA CAGR:
{format_percent(ebitda_cagr)}

PAT CAGR:
{format_percent(pat_cagr)}

FY25 Revenue Growth:
{format_percent(fy25_revenue_growth)}

FY25 EBITDA Growth:
{format_percent(fy25_ebitda_growth)}

FY25 PAT Growth:
{format_percent(fy25_pat_growth)}

Cash Conversion Cycle:
{format_number(ccc)} days


ANNUAL REPORT SOURCES

{source_context}


QUESTION

{question}


RULES

- Use only supplied information.
- Do not invent facts.
- Cite Annual Report evidence as
  [Source 1], [Source 2], etc.
- If information is insufficient,
  clearly say so.
- Do not use LaTeX.
- Write formulas in plain text.
- Keep answer concise.
"""


            client = Groq(
                api_key=api_key
            )


            try:

                answer = (
                    client.chat.completions.create(
                        model="openai/gpt-oss-20b",
                        messages=[
                            {
                                "role": "user",
                                "content": question_prompt
                            }
                        ]
                    )
                )


                st.subheader(
                    "AI Answer"
                )


                with st.container(
                    border=True
                ):

                    st.markdown(
                        answer
                        .choices[0]
                        .message
                        .content
                    )


                if matches:

                    st.subheader(
                        "Source Evidence"
                    )


                    source_columns = st.columns(
                        len(matches)
                    )


                    for index, result in enumerate(
                        matches
                    ):

                        with source_columns[index]:

                            st.html(
                                f"""
                                <div class="source-card">

                                    <b>
                                    Source {index + 1}
                                    </b>

                                    <br><br>

                                    {result['file']}

                                    <br><br>

                                    <b>PDF Page:</b>
                                    {result['page']}

                                </div>
                                """
                            )


                    with st.expander(
                        "View retrieved Annual Report text"
                    ):

                        for index, result in enumerate(
                            matches,
                            start=1
                        ):

                            st.markdown(
                                f"### Source {index}"
                            )

                            st.caption(
                                f"{result['file']} · "
                                f"PDF Page {result['page']}"
                            )

                            st.write(
                                result["text"][:1200]
                            )


                else:

                    st.info(
                        "No matching Annual Report page was found."
                    )


            except Exception as error:

                st.error(
                    f"AI request failed: {error}"
                )


# =========================================================
# TAB 4 — RISK & OPPORTUNITY
# =========================================================

with tab4:

    st.html(
        """
        <div class="section">

            <div class="eyebrow">
                Financial Signals
            </div>

            <div class="section-title">
                AI Risk & Opportunity Detector
            </div>

            <div class="section-desc">
                Identifies financial strengths,
                concerns and positive signals.
            </div>

        </div>
        """
    )


    r1, r2, r3, r4 = st.columns(4)


    r1.metric(
        "Revenue CAGR",
        format_percent(
            revenue_cagr
        )
    )


    r2.metric(
        "EBITDA CAGR",
        format_percent(
            ebitda_cagr
        )
    )


    r3.metric(
        "ROCE",
        format_percent(
            roce,
            1
        )
    )


    r4.metric(
        "Cash Conversion Cycle",
        f"{format_number(ccc)} Days"
    )


    st.write("")


    if st.button(
        "Detect Risks & Opportunities",
        key="risk_button"
    ):

        if not api_key:

            st.error(
                "Groq API key not found."
            )

        else:

            risk_prompt = f"""
You are a professional financial analyst.

Review these Hindustan Unilever financial metrics.

Revenue CAGR:
{format_percent(revenue_cagr)}

EBITDA CAGR:
{format_percent(ebitda_cagr)}

PAT CAGR:
{format_percent(pat_cagr)}

FY25 Revenue Growth:
{format_percent(fy25_revenue_growth)}

FY25 EBITDA Growth:
{format_percent(fy25_ebitda_growth)}

FY25 PAT Growth:
{format_percent(fy25_pat_growth)}

EBITDA Margin:
{format_percent(ebitda_margin)}

ROE:
{format_percent(roe)}

ROCE:
{format_percent(roce, 1)}

Cash Conversion Cycle:
{format_number(ccc)} days


Give exactly:

### Key Strength
One concise point.

### Key Concern
One concise point.

### Opportunity / Positive Signal
One concise point.


Rules:

- Use only supplied numbers.
- Do not invent business causes.
- Do not use LaTeX.
"""


            client = Groq(
                api_key=api_key
            )


            try:

                with st.spinner(
                    "Analysing financial signals..."
                ):

                    result = (
                        client.chat.completions.create(
                            model="openai/gpt-oss-20b",
                            messages=[
                                {
                                    "role": "user",
                                    "content": risk_prompt
                                }
                            ]
                        )
                    )


                with st.container(
                    border=True
                ):

                    st.markdown(
                        result
                        .choices[0]
                        .message
                        .content
                    )


            except Exception as error:

                st.error(
                    f"AI request failed: {error}"
                )


# =========================================================
# TAB 5 — FORECAST
# =========================================================

with tab5:

    st.html(
        """
        <div class="section">

            <div class="eyebrow">
                Forward-Looking Analysis
            </div>

            <div class="section-title">
                Bear / Base / Bull Forecast
            </div>

            <div class="section-desc">
                Compare FY26-FY28 outcomes under
                different financial assumptions.
            </div>

        </div>
        """
    )


    st.html(
        """
        <div class="note">
            Forecasts are scenario-based analytical estimates,
            not guaranteed predictions.
        </div>
        """
    )


    bear_col, base_col, bull_col = st.columns(3)


    with bear_col:

        with st.container(border=True):

            st.subheader("Bear Case")

            bear_growth = st.number_input(
                "Revenue Growth (%)",
                -10.0,
                20.0,
                2.0,
                0.5,
                key="bear_growth"
            )

            bear_ebitda = st.number_input(
                "EBITDA Margin (%)",
                10.0,
                35.0,
                22.0,
                0.5,
                key="bear_ebitda"
            )

            bear_pat = st.number_input(
                "PAT Margin (%)",
                5.0,
                30.0,
                16.0,
                0.5,
                key="bear_pat"
            )


    with base_col:

        with st.container(border=True):

            st.subheader("Base Case")

            base_growth = st.number_input(
                "Revenue Growth (%)",
                -10.0,
                20.0,
                5.0,
                0.5,
                key="base_growth"
            )

            base_ebitda = st.number_input(
                "EBITDA Margin (%)",
                10.0,
                35.0,
                24.0,
                0.5,
                key="base_ebitda"
            )

            base_pat = st.number_input(
                "PAT Margin (%)",
                5.0,
                30.0,
                17.5,
                0.5,
                key="base_pat"
            )


    with bull_col:

        with st.container(border=True):

            st.subheader("Bull Case")

            bull_growth = st.number_input(
                "Revenue Growth (%)",
                -10.0,
                20.0,
                8.0,
                0.5,
                key="bull_growth"
            )

            bull_ebitda = st.number_input(
                "EBITDA Margin (%)",
                10.0,
                35.0,
                25.0,
                0.5,
                key="bull_ebitda"
            )

            bull_pat = st.number_input(
                "PAT Margin (%)",
                5.0,
                30.0,
                18.5,
                0.5,
                key="bull_pat"
            )


    bear_forecast = create_forecast(
        revenue,
        bear_growth,
        bear_ebitda,
        bear_pat
    )


    base_forecast = create_forecast(
        revenue,
        base_growth,
        base_ebitda,
        base_pat
    )


    bull_forecast = create_forecast(
        revenue,
        bull_growth,
        bull_ebitda,
        bull_pat
    )


    st.subheader(
        "FY26 Snapshot"
    )


    f1, f2, f3 = st.columns(3)


    f1.metric(
        "Bear Revenue",
        format_money(
            bear_forecast.iloc[0]["Revenue"]
        )
    )


    f2.metric(
        "Base Revenue",
        format_money(
            base_forecast.iloc[0]["Revenue"]
        )
    )


    f3.metric(
        "Bull Revenue",
        format_money(
            bull_forecast.iloc[0]["Revenue"]
        )
    )


    revenue_fig = go.Figure()


    for scenario, values, color in [
        (
            "Bear Case",
            bear_forecast["Revenue"],
            "#94A3B8"
        ),
        (
            "Base Case",
            base_forecast["Revenue"],
            PRIMARY
        ),
        (
            "Bull Case",
            bull_forecast["Revenue"],
            NAVY
        )
    ]:

        revenue_fig.add_trace(
            go.Scatter(
                x=values.index,
                y=values.values,
                mode="lines+markers",
                name=scenario,
                line=dict(
                    color=color,
                    width=3
                ),
                marker=dict(size=8),
                hovertemplate=(
                    f"<b>{scenario}</b><br>"
                    "%{x}<br>"
                    "₹%{y:,.0f} Cr"
                    "<extra></extra>"
                )
            )
        )


    revenue_fig.update_layout(
        title="Revenue Scenario Comparison",
        yaxis_title="₹ Crore"
    )


    style_chart(
        revenue_fig
    )


    st.plotly_chart(
        revenue_fig,
        use_container_width=True,
        config={
            "displayModeBar": False
        }
    )


    pat_fig = go.Figure()


    for scenario, values, color in [
        (
            "Bear Case",
            bear_forecast["PAT"],
            "#94A3B8"
        ),
        (
            "Base Case",
            base_forecast["PAT"],
            PRIMARY
        ),
        (
            "Bull Case",
            bull_forecast["PAT"],
            NAVY
        )
    ]:

        pat_fig.add_trace(
            go.Bar(
                x=values.index,
                y=values.values,
                name=scenario,
                marker_color=color,
                hovertemplate=(
                    f"<b>{scenario}</b><br>"
                    "%{x}<br>"
                    "₹%{y:,.0f} Cr"
                    "<extra></extra>"
                )
            )
        )


    pat_fig.update_layout(
        title="PAT Scenario Comparison",
        yaxis_title="₹ Crore",
        barmode="group"
    )


    style_chart(
        pat_fig
    )


    st.plotly_chart(
        pat_fig,
        use_container_width=True,
        config={
            "displayModeBar": False
        }
    )


    st.subheader(
        "Detailed Forecast"
    )


    selected_scenario = st.selectbox(
        "Select scenario",
        [
            "Bear Case",
            "Base Case",
            "Bull Case"
        ],
        key="forecast_scenario"
    )


    if selected_scenario == "Bear Case":

        selected_forecast = bear_forecast

    elif selected_scenario == "Base Case":

        selected_forecast = base_forecast

    else:

        selected_forecast = bull_forecast


    display_forecast = selected_forecast.copy()


    for column in [
        "Revenue",
        "EBITDA",
        "PAT"
    ]:

        display_forecast[column] = (
            display_forecast[column]
            .apply(
                lambda value:
                f"₹{value:,.0f} Cr"
            )
        )


    st.dataframe(
        display_forecast,
        use_container_width=True
    )


    st.subheader(
        "AI Scenario Interpretation"
    )


    if st.button(
        "Compare Scenarios with AI",
        key="scenario_ai"
    ):

        if not api_key:

            st.error(
                "Groq API key not found."
            )

        else:

            scenario_prompt = f"""
You are a professional financial analyst.

Compare these HUL scenarios.

BEAR CASE
Growth: {bear_growth:.1f}%
EBITDA Margin: {bear_ebitda:.1f}%
PAT Margin: {bear_pat:.1f}%
FY28 Revenue:
{format_money(bear_forecast.iloc[2]["Revenue"])}
FY28 PAT:
{format_money(bear_forecast.iloc[2]["PAT"])}

BASE CASE
Growth: {base_growth:.1f}%
EBITDA Margin: {base_ebitda:.1f}%
PAT Margin: {base_pat:.1f}%
FY28 Revenue:
{format_money(base_forecast.iloc[2]["Revenue"])}
FY28 PAT:
{format_money(base_forecast.iloc[2]["PAT"])}

BULL CASE
Growth: {bull_growth:.1f}%
EBITDA Margin: {bull_ebitda:.1f}%
PAT Margin: {bull_pat:.1f}%
FY28 Revenue:
{format_money(bull_forecast.iloc[2]["Revenue"])}
FY28 PAT:
{format_money(bull_forecast.iloc[2]["PAT"])}


Give:

### Bear Case
### Base Case
### Bull Case
### Scenario Comparison
### Key Sensitivity


Rules:

- Treat figures as scenario estimates.
- Use only supplied numbers.
- Do not invent events.
- Do not use LaTeX.
"""


            client = Groq(
                api_key=api_key
            )


            try:

                with st.spinner(
                    "Comparing scenarios..."
                ):

                    response = (
                        client.chat.completions.create(
                            model="openai/gpt-oss-20b",
                            messages=[
                                {
                                    "role": "user",
                                    "content": scenario_prompt
                                }
                            ]
                        )
                    )


                with st.container(
                    border=True
                ):

                    st.markdown(
                        response
                        .choices[0]
                        .message
                        .content
                    )


            except Exception as error:

                st.error(
                    f"AI request failed: {error}"
                )


# =========================================================
# TAB 6 — SENSITIVITY
# =========================================================

with tab6:

    st.html(
        """
        <div class="section">

            <div class="eyebrow">
                Financial Sensitivity
            </div>

            <div class="section-title">
                Growth & Margin Sensitivity
            </div>

            <div class="section-desc">
                Analyse how Revenue Growth and margin
                changes affect FY26 EBITDA and PAT.
            </div>

        </div>
        """
    )


    st.html(
        """
        <div class="note">
            Heatmaps show how financial outcomes
            respond to different assumptions.
        </div>
        """
    )


    s1, s2 = st.columns(2)


    with s1:

        selected_growth = st.slider(
            "Base Revenue Growth (%)",
            0.0,
            10.0,
            5.0,
            0.5,
            key="sensitivity_growth"
        )


    with s2:

        selected_margin = st.slider(
            "Base EBITDA Margin (%)",
            20.0,
            28.0,
            24.0,
            0.5,
            key="sensitivity_margin"
        )


    selected_revenue = (
        revenue
        * (
            1
            + selected_growth / 100
        )
    )


    selected_ebitda = (
        selected_revenue
        * selected_margin
        / 100
    )


    sm1, sm2 = st.columns(2)


    sm1.metric(
        "Selected FY26 Revenue",
        format_money(
            selected_revenue
        )
    )


    sm2.metric(
        "Selected FY26 EBITDA",
        format_money(
            selected_ebitda
        )
    )


    growth_rates = [
        2.0,
        4.0,
        5.0,
        6.0,
        8.0
    ]


    ebitda_margins = [
        22.0,
        23.0,
        24.0,
        25.0,
        26.0
    ]


    ebitda_matrix = []


    for growth_rate in growth_rates:

        future_revenue = (
            revenue
            * (
                1
                + growth_rate / 100
            )
        )

        row = []

        for margin in ebitda_margins:

            row.append(
                future_revenue
                * margin
                / 100
            )

        ebitda_matrix.append(row)


    ebitda_df = pd.DataFrame(
        ebitda_matrix,
        index=[
            f"{x:.0f}%"
            for x in growth_rates
        ],
        columns=[
            f"{x:.0f}%"
            for x in ebitda_margins
        ]
    )


    st.subheader(
        "FY26 EBITDA Sensitivity"
    )


    st.caption(
        "Rows = Revenue Growth | Columns = EBITDA Margin"
    )


    ebitda_heatmap = px.imshow(
        ebitda_df,
        text_auto=".0f",
        aspect="auto",
        labels={
            "x": "EBITDA Margin",
            "y": "Revenue Growth",
            "color": "EBITDA"
        },
        color_continuous_scale=[
            [0, "#EFF6FF"],
            [1, PRIMARY]
        ]
    )


    ebitda_heatmap.update_traces(
        hovertemplate=(
            "Revenue Growth: %{y}<br>"
            "EBITDA Margin: %{x}<br>"
            "FY26 EBITDA: ₹%{z:,.0f} Cr"
            "<extra></extra>"
        )
    )


    ebitda_heatmap.update_layout(
        height=410,
        paper_bgcolor=WHITE,
        plot_bgcolor=WHITE,
        margin=dict(
            l=35,
            r=35,
            t=25,
            b=35
        ),
        coloraxis_colorbar=dict(
            title="₹ Cr"
        )
    )


    st.plotly_chart(
        ebitda_heatmap,
        use_container_width=True,
        config={
            "displayModeBar": False
        }
    )


    pat_margins = [
        15.0,
        16.0,
        17.0,
        18.0,
        19.0
    ]


    pat_matrix = []


    for growth_rate in growth_rates:

        future_revenue = (
            revenue
            * (
                1
                + growth_rate / 100
            )
        )

        row = []

        for margin in pat_margins:

            row.append(
                future_revenue
                * margin
                / 100
            )

        pat_matrix.append(row)


    pat_df = pd.DataFrame(
        pat_matrix,
        index=[
            f"{x:.0f}%"
            for x in growth_rates
        ],
        columns=[
            f"{x:.0f}%"
            for x in pat_margins
        ]
    )


    st.subheader(
        "FY26 PAT Sensitivity"
    )


    st.caption(
        "Rows = Revenue Growth | Columns = PAT Margin"
    )


    pat_heatmap = px.imshow(
        pat_df,
        text_auto=".0f",
        aspect="auto",
        labels={
            "x": "PAT Margin",
            "y": "Revenue Growth",
            "color": "PAT"
        },
        color_continuous_scale=[
            [0, "#EFF6FF"],
            [1, PRIMARY]
        ]
    )


    pat_heatmap.update_traces(
        hovertemplate=(
            "Revenue Growth: %{y}<br>"
            "PAT Margin: %{x}<br>"
            "FY26 PAT: ₹%{z:,.0f} Cr"
            "<extra></extra>"
        )
    )


    pat_heatmap.update_layout(
        height=410,
        paper_bgcolor=WHITE,
        plot_bgcolor=WHITE,
        margin=dict(
            l=35,
            r=35,
            t=25,
            b=35
        ),
        coloraxis_colorbar=dict(
            title="₹ Cr"
        )
    )


    st.plotly_chart(
        pat_heatmap,
        use_container_width=True,
        config={
            "displayModeBar": False
        }
    )


    minimum_ebitda = (
        ebitda_df.min().min()
    )

    maximum_ebitda = (
        ebitda_df.max().max()
    )

    minimum_pat = (
        pat_df.min().min()
    )

    maximum_pat = (
        pat_df.max().max()
    )


    st.subheader(
        "Outcome Range"
    )


    o1, o2, o3, o4 = st.columns(4)


    o1.metric(
        "Lowest EBITDA",
        format_money(
            minimum_ebitda
        )
    )


    o2.metric(
        "Highest EBITDA",
        format_money(
            maximum_ebitda
        )
    )


    o3.metric(
        "Lowest PAT",
        format_money(
            minimum_pat
        )
    )


    o4.metric(
        "Highest PAT",
        format_money(
            maximum_pat
        )
    )


    st.subheader(
        "AI Sensitivity Interpretation"
    )


    if st.button(
        "Explain Sensitivity with AI",
        key="sensitivity_ai"
    ):

        if not api_key:

            st.error(
                "Groq API key not found."
            )

        else:

            sensitivity_prompt = f"""
You are a professional financial analyst.

Interpret this HUL FY26 sensitivity analysis.

FY25 Revenue:
{format_money(revenue)}

Revenue Growth Tested:
2%, 4%, 5%, 6%, 8%

EBITDA Margins Tested:
22%, 23%, 24%, 25%, 26%

PAT Margins Tested:
15%, 16%, 17%, 18%, 19%

Lowest EBITDA:
{format_money(minimum_ebitda)}

Highest EBITDA:
{format_money(maximum_ebitda)}

Lowest PAT:
{format_money(minimum_pat)}

Highest PAT:
{format_money(maximum_pat)}


Give:

### EBITDA Sensitivity
### PAT Sensitivity
### Key Driver
### Financial Planning Insight


Rules:

- Use only supplied information.
- Do not invent causes.
- Treat outputs as sensitivity estimates.
- Do not use LaTeX.
"""


            client = Groq(
                api_key=api_key
            )


            try:

                with st.spinner(
                    "Analysing sensitivity..."
                ):

                    response = (
                        client.chat.completions.create(
                            model="openai/gpt-oss-20b",
                            messages=[
                                {
                                    "role": "user",
                                    "content": sensitivity_prompt
                                }
                            ]
                        )
                    )


                with st.container(
                    border=True
                ):

                    st.markdown(
                        response
                        .choices[0]
                        .message
                        .content
                    )


            except Exception as error:

                st.error(
                    f"AI request failed: {error}"
                )


# =========================================================
# TAB 7 — VALUATION
# =========================================================

with tab7:

    st.html(
        """
        <div class="section">

            <div class="eyebrow">
                Illustrative Valuation
            </div>

            <div class="section-title">
                Multiples-Based Valuation
            </div>

            <div class="section-desc">
                Estimate HUL's implied value using
                P/E and EV/EBITDA valuation approaches.
            </div>

        </div>
        """
    )


    st.html(
        """
        <div class="note">

            Multiples below are user-selected assumptions.
            This module is designed for scenario analysis,
            not as a current market-price recommendation.

        </div>
        """
    )


    # -----------------------------------------------------
    # ACTUAL FINANCIAL BASE
    # -----------------------------------------------------

    st.subheader(
        "FY25 Valuation Base"
    )


    v1, v2, v3, v4 = st.columns(4)


    v1.metric(
        "FY25 PAT",
        format_money(pat)
    )


    v2.metric(
        "FY25 EBITDA",
        format_money(ebitda)
    )


    v3.metric(
        "Cash",
        format_money(cash)
    )


    v4.metric(
        "Lease Liabilities",
        format_money(
            lease_liabilities
        )
    )


    # -----------------------------------------------------
    # ASSUMPTIONS
    # -----------------------------------------------------

    st.subheader(
        "Valuation Assumptions"
    )


    val_col1, val_col2 = st.columns(2)


    with val_col1:

        pe_multiple = st.slider(
            "P/E Multiple (x)",
            min_value=20.0,
            max_value=80.0,
            value=50.0,
            step=1.0,
            key="pe_multiple"
        )


    with val_col2:

        ev_ebitda_multiple = st.slider(
            "EV / EBITDA Multiple (x)",
            min_value=10.0,
            max_value=50.0,
            value=30.0,
            step=1.0,
            key="ev_ebitda_multiple"
        )


    # -----------------------------------------------------
    # P/E VALUATION
    # -----------------------------------------------------

    pe_equity_value = (
        pat
        * pe_multiple
    )


    # -----------------------------------------------------
    # EV / EBITDA VALUATION
    # -----------------------------------------------------

    enterprise_value = (
        ebitda
        * ev_ebitda_multiple
    )


    ev_equity_value = (
        enterprise_value
        + cash
        - lease_liabilities
    )


    # -----------------------------------------------------
    # RESULTS
    # -----------------------------------------------------

    st.subheader(
        "Implied Valuation"
    )


    value1, value2 = st.columns(2)


    with value1:

        st.metric(
            "P/E Implied Equity Value",
            format_money(
                pe_equity_value
            )
        )

        st.caption(
            f"PAT × {pe_multiple:.0f}x P/E"
        )


    with value2:

        st.metric(
            "EV/EBITDA Implied Equity Value",
            format_money(
                ev_equity_value
            )
        )

        st.caption(
            (
                f"EBITDA × "
                f"{ev_ebitda_multiple:.0f}x "
                f"+ Cash - Lease Liabilities"
            )
        )


    st.html(
        """
        <div class="note">

            <b>P/E approach:</b>
            Equity Value = PAT × P/E Multiple

            <br>

            <b>EV/EBITDA approach:</b>
            Enterprise Value = EBITDA × EV/EBITDA Multiple

            <br>

            Approximate Equity Value =
            Enterprise Value + Cash - Lease Liabilities

        </div>
        """
    )


    # -----------------------------------------------------
    # COMPARISON CHART
    # -----------------------------------------------------

    valuation_fig = go.Figure()


    valuation_fig.add_trace(
        go.Bar(
            x=[
                "P/E Method",
                "EV/EBITDA Method"
            ],
            y=[
                pe_equity_value,
                ev_equity_value
            ],
            marker_color=[
                PRIMARY,
                NAVY
            ],
            text=[
                f"₹{pe_equity_value:,.0f} Cr",
                f"₹{ev_equity_value:,.0f} Cr"
            ],
            textposition="outside",
            hovertemplate=(
                "%{x}<br>"
                "₹%{y:,.0f} Cr"
                "<extra></extra>"
            )
        )
    )


    valuation_fig.update_layout(
        title="Implied Equity Value Comparison",
        yaxis_title="₹ Crore",
        showlegend=False
    )


    style_chart(
        valuation_fig
    )


    st.plotly_chart(
        valuation_fig,
        use_container_width=True,
        config={
            "displayModeBar": False
        }
    )


    # -----------------------------------------------------
    # P/E SENSITIVITY
    # -----------------------------------------------------

    st.subheader(
        "P/E Valuation Sensitivity"
    )


    pe_multiples = [
        35,
        40,
        45,
        50,
        55,
        60,
        65
    ]


    pat_cases = [
        pat * 0.90,
        pat * 0.95,
        pat,
        pat * 1.05,
        pat * 1.10
    ]


    pat_case_labels = [
        "-10%",
        "-5%",
        "Base PAT",
        "+5%",
        "+10%"
    ]


    pe_matrix = []


    for pat_case in pat_cases:

        row = []

        for multiple in pe_multiples:

            row.append(
                pat_case
                * multiple
            )

        pe_matrix.append(row)


    pe_sensitivity_df = pd.DataFrame(
        pe_matrix,
        index=pat_case_labels,
        columns=[
            f"{x}x"
            for x in pe_multiples
        ]
    )


    st.caption(
        "Rows = PAT variation | Columns = P/E multiple"
    )


    pe_heatmap = px.imshow(
        pe_sensitivity_df,
        text_auto=".0f",
        aspect="auto",
        labels={
            "x": "P/E Multiple",
            "y": "PAT Scenario",
            "color": "Equity Value"
        },
        color_continuous_scale=[
            [0, "#EFF6FF"],
            [1, PRIMARY]
        ]
    )


    pe_heatmap.update_traces(
        hovertemplate=(
            "PAT Scenario: %{y}<br>"
            "P/E Multiple: %{x}<br>"
            "Equity Value: ₹%{z:,.0f} Cr"
            "<extra></extra>"
        )
    )


    pe_heatmap.update_layout(
        height=420,
        paper_bgcolor=WHITE,
        plot_bgcolor=WHITE,
        margin=dict(
            l=35,
            r=35,
            t=25,
            b=35
        ),
        coloraxis_colorbar=dict(
            title="₹ Cr"
        )
    )


    st.plotly_chart(
        pe_heatmap,
        use_container_width=True,
        config={
            "displayModeBar": False
        }
    )


    # -----------------------------------------------------
    # EV / EBITDA SENSITIVITY
    # -----------------------------------------------------

    st.subheader(
        "EV / EBITDA Valuation Sensitivity"
    )


    ev_multiples = [
        20,
        24,
        28,
        30,
        32,
        36,
        40
    ]


    ebitda_cases = [
        ebitda * 0.90,
        ebitda * 0.95,
        ebitda,
        ebitda * 1.05,
        ebitda * 1.10
    ]


    ebitda_case_labels = [
        "-10%",
        "-5%",
        "Base EBITDA",
        "+5%",
        "+10%"
    ]


    ev_matrix = []


    for ebitda_case in ebitda_cases:

        row = []

        for multiple in ev_multiples:

            ev = (
                ebitda_case
                * multiple
            )

            equity = (
                ev
                + cash
                - lease_liabilities
            )

            row.append(
                equity
            )

        ev_matrix.append(row)


    ev_sensitivity_df = pd.DataFrame(
        ev_matrix,
        index=ebitda_case_labels,
        columns=[
            f"{x}x"
            for x in ev_multiples
        ]
    )


    st.caption(
        "Rows = EBITDA variation | Columns = EV/EBITDA multiple"
    )


    ev_heatmap = px.imshow(
        ev_sensitivity_df,
        text_auto=".0f",
        aspect="auto",
        labels={
            "x": "EV/EBITDA Multiple",
            "y": "EBITDA Scenario",
            "color": "Equity Value"
        },
        color_continuous_scale=[
            [0, "#EFF6FF"],
            [1, PRIMARY]
        ]
    )


    ev_heatmap.update_traces(
        hovertemplate=(
            "EBITDA Scenario: %{y}<br>"
            "EV/EBITDA Multiple: %{x}<br>"
            "Equity Value: ₹%{z:,.0f} Cr"
            "<extra></extra>"
        )
    )


    ev_heatmap.update_layout(
        height=420,
        paper_bgcolor=WHITE,
        plot_bgcolor=WHITE,
        margin=dict(
            l=35,
            r=35,
            t=25,
            b=35
        ),
        coloraxis_colorbar=dict(
            title="₹ Cr"
        )
    )


    st.plotly_chart(
        ev_heatmap,
        use_container_width=True,
        config={
            "displayModeBar": False
        }
    )


    # -----------------------------------------------------
    # AI VALUATION
    # -----------------------------------------------------

    st.subheader(
        "AI Valuation Interpretation"
    )


    if st.button(
        "Explain Valuation with AI",
        key="valuation_ai"
    ):

        if not api_key:

            st.error(
                "Groq API key not found."
            )

        else:

            valuation_prompt = f"""
You are a professional financial analyst.

Interpret the following illustrative
Hindustan Unilever valuation.

FY25 PAT:
{format_money(pat)}

FY25 EBITDA:
{format_money(ebitda)}

Cash:
{format_money(cash)}

Lease Liabilities:
{format_money(lease_liabilities)}


P/E ASSUMPTION

P/E Multiple:
{pe_multiple:.0f}x

P/E Implied Equity Value:
{format_money(pe_equity_value)}


EV/EBITDA ASSUMPTION

EV/EBITDA Multiple:
{ev_ebitda_multiple:.0f}x

Enterprise Value:
{format_money(enterprise_value)}

EV/EBITDA Implied Equity Value:
{format_money(ev_equity_value)}


Give exactly:

### P/E Valuation
Explain the P/E result.

### EV/EBITDA Valuation
Explain the EV/EBITDA result.

### Method Comparison
Compare the two implied equity values.

### Key Valuation Sensitivity
Explain why changes in earnings,
EBITDA or multiples affect valuation.

### Important Limitation
State that valuation depends heavily
on selected multiples and assumptions.


Rules:

- Use only supplied figures.
- Do not invent market-price data.
- Do not say that the stock is
  undervalued or overvalued.
- Do not provide investment advice.
- Do not use LaTeX.
- Keep response concise.
"""


            client = Groq(
                api_key=api_key
            )


            try:

                with st.spinner(
                    "Analysing valuation..."
                ):

                    response = (
                        client.chat.completions.create(
                            model="openai/gpt-oss-20b",
                            messages=[
                                {
                                    "role": "user",
                                    "content": valuation_prompt
                                }
                            ]
                        )
                    )


                with st.container(
                    border=True
                ):

                    st.markdown(
                        response
                        .choices[0]
                        .message
                        .content
                    )


            except Exception as error:

                st.error(
                    f"AI request failed: {error}"
                )


    st.caption(
        "Valuation outputs are illustrative scenario estimates "
        "based on user-selected multiples. They are not investment advice."
    )


# =========================================================
# FOOTER
# =========================================================

st.html(
    """
    <div class="footer-box">

        <b>Project:</b>
        AI-Powered Financial Analysis —
        Hindustan Unilever Limited

        <br>

        <b>Financial Data:</b>
        Standalone financial statements,
        FY2020-21 to FY2024-25

        <br>

        <b>Sources:</b>
        HUL Annual Reports and company-reported
        financial information

        <br>

        <b>Technology:</b>
        Excel · Python · Streamlit · Plotly ·
        Groq LLM · PDF Retrieval

        <br>

        <b>Modules:</b>
        Historical Analysis · AI Analysis ·
        Annual Report Q&A · Risk Analysis ·
        Forecasting · Sensitivity Analysis ·
        Valuation

        <br>

        <b>Disclaimer:</b>
        Forecast, sensitivity and valuation outputs
        are scenario-based analytical estimates and
        are not investment advice.

    </div>
    """
)